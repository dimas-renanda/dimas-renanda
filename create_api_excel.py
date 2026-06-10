import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# ============== SHEET 1: callApiFetch ==============
ws1 = wb.active
ws1.title = "callApiFetch"

# Headers
headers1 = ["No", "File", "Line", "Function", "modulCode", "Parameters", "Data Type", "Description"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Data callApiFetch
callapi_data = [
    # TRUCK
    [1, "lib/controller/trucklist_c.dart", 63, "fetchTrucks(BuildContext context)", "getTruck", "-", "-", "Fetch all trucks"],
    [2, "lib/controller/trucklist_c.dart", 84, "resetNfc(String truckId)", "resetNfc", "id", "String", "Reset NFC for truck"],
    [3, "lib/controller/trucklist_c.dart", 105, "changeTruck(String truckId, String set)", "saveTruckAvailable", "id, available", "String, String", "Enable/disable truck"],
    [4, "lib/controller/trucklist_c.dart", 130, "fetchLocations()", "getLokasi", "-", "-", "Fetch locations"],
    [5, "lib/controller/trucklist_c.dart", 154, "changeLocation(String truckId, String lokasi)", "saveTruckLokasi", "id, lokasi", "String, String", "Save truck location"],
    [6, "lib/controller/trucklist_c.dart", 297, "sendNFCTags(String nfctags, String truckId, String platNo)", "saveNfcTruck", "id, nfc, platNo", "String, String, String", "Save NFC to truck"],
    
    # INVENTORY
    [7, "lib/controller/inventorylist_c.dart", 92, "fetchInventory(BuildContext context)", "invGetItemStockOpname", "-", "-", "Fetch inventory"],
    [8, "lib/controller/inventorylist_c.dart", 115, "resetNfc(String inventoryId)", "resetNfc", "id", "String", "Reset NFC for inventory"],
    [9, "lib/controller/inventorylist_c.dart", 137, "changeInventory(String inventoryId, String set)", "saveInventoryAvailable", "id, available", "String, String", "Enable/disable inventory"],
    [10, "lib/controller/inventorylist_c.dart", 162, "fetchLocations()", "getLokasi", "-", "-", "Fetch locations"],
    [11, "lib/controller/inventorylist_c.dart", 183, "changeLocation(String inventoryId, String lokasi)", "saveInventoryLokasi", "id, lokasi", "String, String", "Save inventory location"],
    [12, "lib/controller/inventorylist_c.dart", 311, "sendNFCTags(String nfctags, String inventoryId, String inventoryNo)", "saveNfcInventory", "id, nfc, inventoryNo", "String, String, String", "Save NFC to inventory"],
    [13, "lib/controller/inventorylist_c.dart", 366, "deleteInventory(String id, BuildContext context)", "invDeleteItemStockOpname", "id", "String", "Delete inventory item"],
    
    # INVENTORY DETAIL
    [14, "lib/controller/inventory_detail_controller.dart", 18, "fetchSites()", "invGetSite", "-", "-", "Fetch sites"],
    [15, "lib/controller/inventory_detail_controller.dart", 28, "fetchItems()", "invGetItem", "-", "-", "Fetch items"],
    [16, "lib/controller/inventory_detail_controller.dart", 48, "saveStockOpname(dynamic body)", "invSaveItemStockOpname", "body", "Map<String, dynamic>", "Save stock opname"],
    
    # GENSET
    [17, "lib/controller/gensetlist_c.dart", 63, "fetchGensets(BuildContext context)", "getGenset", "-", "-", "Fetch all gensets"],
    [18, "lib/controller/gensetlist_c.dart", 84, "resetNfc(String gensetId)", "resetNfc", "id", "String", "Reset NFC for genset"],
    [19, "lib/controller/gensetlist_c.dart", 105, "changeGenset(String gensetId, String set)", "saveGensetAvailable", "id, available", "String, String", "Enable/disable genset"],
    [20, "lib/controller/gensetlist_c.dart", 129, "fetchLocations()", "getLokasi", "-", "-", "Fetch locations"],
    [21, "lib/controller/gensetlist_c.dart", 150, "changeLocation(String gensetId, String lokasi)", "saveGensetLokasi", "id, lokasi", "String, String", "Save genset location"],
    [22, "lib/controller/gensetlist_c.dart", 276, "sendNFCTags(String nfctags, String gensetId, String gensetNo)", "saveNfcGenset", "id, nfc, gensetNo", "String, String, String", "Save NFC to genset"],
    
    # SASIS
    [23, "lib/controller/sasislist_c.dart", 63, "fetchSasis(BuildContext context)", "getSasis", "-", "-", "Fetch all sasis"],
    [24, "lib/controller/sasislist_c.dart", 84, "resetNfc(String sasisId)", "resetNfc", "id", "String", "Reset NFC for sasis"],
    [25, "lib/controller/sasislist_c.dart", 105, "changeSasis(String sasisId, String set)", "saveSasisAvailable", "id, available", "String, String", "Enable/disable sasis"],
    [26, "lib/controller/sasislist_c.dart", 129, "fetchLocations()", "getLokasi", "-", "-", "Fetch locations"],
    [27, "lib/controller/sasislist_c.dart", 150, "changeLocation(String sasisId, String lokasi)", "saveSasisLokasi", "id, lokasi", "String, String", "Save sasis location"],
    [28, "lib/controller/sasislist_c.dart", 276, "sendNFCTags(String nfctags, String sasisId, String sasisNo)", "saveNfcSasis", "id, nfc, sasisNo", "String, String, String", "Save NFC to sasis"],
    
    # ASSET
    [29, "lib/controller/assetlist_c.dart", 77, "fetchAssets(BuildContext context)", "getAsset", "-", "-", "Fetch all assets"],
    [30, "lib/controller/assetlist_c.dart", 190, "_sendNFCTags(String nfctags, String assetId, String assetName)", "saveNfcAsset", "id, nfc, assetName", "String, String, String", "Save NFC to asset"],
    
    # DRIVER
    [31, "lib/controller/driver_list_controller.dart", 10, "fetchDrivers()", "getDriver", "-", "-", "Fetch all drivers"],
    [32, "lib/controller/driver_list_controller.dart", 33, "setDriverEnabled(int driverId, bool enabled)", "saveDriverEnabled", "id, enabled", "int, bool", "Enable/disable driver"],
    [33, "lib/controller/driver_list_controller.dart", 55, "setDriverStatus(int driverId, String status)", "saveDriverStatus", "id, status", "int, String", "Set driver status"],
    
    # DRIVER ORDER
    [34, "lib/controller/driver_order_controller.dart", 26, "fetchOrders()", "toGetDriverJob", "-", "-", "Fetch driver jobs"],
    [35, "lib/controller/driver_order_controller.dart", 179, "nfcplatno(String id, String rfid)", "toSaveDriverJobTruckPlatNo", "id, platNo", "String, String", "Save truck plat no via NFC"],
    [36, "lib/controller/driver_order_controller.dart", 203, "nfcsasisno(String id, String rfid)", "toSaveDriverJobSasisNo", "id, sasisNo", "String, String", "Save sasis no via NFC"],
    [37, "lib/controller/driver_order_controller.dart", 227, "nfcgensetno(String id, String rfid)", "toSaveDriverJobGensetNo", "id, gensetNo", "String, String", "Save genset no via NFC"],
    
    # ADMIN TRUCK DALAM
    [38, "lib/controller/admin_truck_dalam_controller.dart", 30, "callapifetch(String modulcode)", "Dynamic", "-", "-", "Generic API call"],
    [39, "lib/controller/admin_truck_dalam_controller.dart", 235, "updateDriver(String toId, [String? driverId])", "toSaveDriver", "id, driverId", "String, String?", "Update driver dalam"],
    [40, "lib/controller/admin_truck_dalam_controller.dart", 244, "cancelTO(String toId, String trxType)", "toSaveTrxType", "id, trxType", "String, String", "Cancel TO"],
    
    # ADMIN TRUCK LUAR
    [41, "lib/controller/admin_truck_luar_controller.dart", 30, "callapifetch(String modulcode)", "Dynamic", "-", "-", "Generic API call"],
    [42, "lib/controller/admin_truck_luar_controller.dart", 230, "updateDriverLuar(String toId, [String? driverId])", "toSaveDriverLuar", "id, driverId", "String, String?", "Update driver luar"],
    [43, "lib/controller/admin_truck_luar_controller.dart", 239, "moveToInside(String toId, String trxType)", "toSaveTrxType", "id, trxType", "String, String", "Move to inside"],
    
    # USER
    [44, "lib/controller/userlistpage_c.dart", 43, "toggleBs(String userId, String currentBs)", "saveUserBs", "userId, bs", "String, String", "Toggle BS"],
    [45, "lib/controller/userlistpage_c.dart", 84, "fetchUsers(BuildContext context)", "getSnituser", "-", "-", "Fetch users"],
    [46, "lib/controller/userlistpage_c.dart", 105, "resetImei(String userId)", "resetImei", "id", "String", "Reset user IMEI"],
    [47, "lib/controller/userlistpage_c.dart", 126, "resetPassword(String userId)", "resetPassword", "id", "String", "Reset user password"],
    [48, "lib/controller/userlistpage_c.dart", 147, "changeUser(String userId, String set)", "saveUserEnabled", "id, enabled", "String, String", "Enable/disable user"],
    [49, "lib/controller/userlistpage_c.dart", 172, "fetchLocations()", "getLokasi", "-", "-", "Fetch locations"],
    [50, "lib/controller/userlistpage_c.dart", 195, "changeLocation(String userId, String lokasi)", "saveSnituserLokasi", "id, lokasi", "String, String", "Save user location"],
    [51, "lib/controller/userlistpage_c.dart", 222, "enableBs(String userId)", "enableBs", "id, bs", "String, String", "Enable BS for user"],
    
    # POM
    [52, "lib/controller/pump_machine_master_c.dart", 63, "fetchMachines(BuildContext context)", "getPom", "-", "-", "Fetch POM machines"],
    [53, "lib/controller/pump_machine_master_c.dart", 85, "changeEnable(String machineId, String set)", "savePomEnabled", "id, enabled", "String, bool", "Enable/disable POM"],
    [54, "lib/controller/pump_machine_master_c.dart", 107, "resetNfc(String machineId)", "resetPumpNfc", "id", "String", "Reset POM NFC"],
    [55, "lib/controller/pump_machine_master_c.dart", 226, "_sendNfcToServer(String nfctags, String machineId, String machineName)", "saveNfcPom", "id, nfc, machineName", "String, String, String", "Save NFC to POM"],
    
    # BS CONTROLLER
    [56, "lib/controller/bs_controller.dart", 45, "onInit()", "bsCheckUser", "-", "-", "Check BS user status"],
    [57, "lib/controller/besverify_c.dart", 107, "fetchUsers()", "bsGetVerify", "-", "-", "Fetch BS verify data"],
    [58, "lib/controller/bs_done_controller.dart", 59, "fetchList()", "bsGetDone", "-", "-", "Fetch BS done list"],
    [59, "lib/controller/bsbypass_c.dart", 75, "bypassUser(String userId)", "bsSavePass", "password", "String", "Save bypass password"],
    [60, "lib/controller/bsbypass_c.dart", 106, "changeUser(String userId, String set)", "saveUserEnabled", "id, enabled", "String, String", "Enable/disable BS user"],
    [61, "lib/controller/bsbypass_c.dart", 138, "fetchUsers()", "bsGetSnituser", "-", "-", "Fetch BS users"],
    [62, "lib/controller/bsbypass_c.dart", 160, "fetchUsersBS()", "bsGetPass", "-", "-", "Fetch BS pass data"],
    [63, "lib/controller/processbspj_controller.dart", 13, "fetchItems()", "bsGetApproval", "-", "-", "Fetch BS approval items"],
    [64, "lib/controller/processbspj_controller.dart", 36, "saveApproval(String id)", "bsSaveApproval", "id, approval", "String, String", "Save BS approval"],
    [65, "lib/controller/processbspj_controller.dart", 58, "saveApproval(String id)", "bsSaveApproval", "id", "String", "Save BS approval (alt)"],
    [66, "lib/controller/repobs_controller.dart", 17, "fetchItems()", "bsGetRepoApproval", "-", "-", "Fetch BS repo approval"],
    
    # DASHBOARD
    [67, "lib/controller/dashboard_location_controller.dart", 16, "fetchLocations()", "dashboardCheckUser", "-", "-", "Check dashboard user"],
    [68, "lib/controller/dashboard_destination_controller.dart", 70, "onInit()", "dashboardIN", "lokasi", "String", "Dashboard IN with location"],
    [69, "lib/controller/dashboard_stripping_controller.dart", 56, "fetchData()", "dashboardStripping", "lokasi", "String?", "Fetch stripping data"],
    [70, "lib/controller/dashboard_stuffing_controller.dart", 56, "fetchData()", "dashboardStuffing", "lokasi", "String?", "Fetch stuffing data"],
    
    # PANTRY
    [71, "lib/controller/pantry_controller.dart", 298, "saveOrder(Map<String, dynamic> orderData)", "pantrySaveOrder", "orderData", "Map<String, dynamic>", "Save pantry order"],
    [72, "lib/controller/pantry_controller.dart", 389, "cancelOrder(String orderId)", "pantryCancelOrder", "orderId", "String", "Cancel pantry order"],
    
    # PATROL
    [73, "lib/controller/patrol_controller.dart", 43, "fetchData()", "patrolGetData", "-", "-", "Fetch patrol data"],
    [74, "lib/controller/patrol_controller.dart", 203, "completeTask(String taskId, DateTime completedAt)", "patrolCompleteTask", "taskId, completedAt", "String, String", "Complete patrol task"],
    
    # HIKVISION
    [75, "lib/controller/hikvision_controller.dart", 128, "createUser(Map<String, dynamic> userData)", "createHikvisionUserRequest", "userData", "Map<String, dynamic>", "Create Hikvision user"],
    [76, "lib/controller/hikvision_controller.dart", 173, "updateUser(Map<String, dynamic> userData)", "updateHikvisionUserRequest", "userData", "Map<String, dynamic>", "Update Hikvision user"],
    [77, "lib/controller/hikvision_controller.dart", 218, "deleteUser(String userId)", "deleteHikvisionUserRequest", "id", "String", "Delete Hikvision user"],
    [78, "lib/controller/hikvision_controller.dart", 271, "openDoor(String deviceId)", "hikvisionDeviceOpenRequest", "deviceId", "String", "Open Hikvision door"],
    [79, "lib/controller/hikvision_controller.dart", 312, "(various)", "hikvisionPersonSetStatusRequest", "-", "-", "Set person status"],
    
    # SALES
    [80, "lib/controller/sales_customer_c.dart", 90, "fetchCustomer(String salescode)", "getCustomer", "sales", "String", "Fetch customers by sales"],
    [81, "lib/controller/salescustomer_c.dart", 90, "fetchCustomer()", "getCustomer", "sales", "String", "Fetch customers"],
]

# Add data to sheet 1
for row_idx, row_data in enumerate(callapi_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws1.cell(row=row_idx, column=col_idx, value=value)

# ============== SHEET 2: callApiFetch View ==============
ws2 = wb.create_sheet("callApiFetch_View")

headers2 = ["No", "File", "Line", "Function", "modulCode", "Parameters", "Data Type", "Description"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

callapi_view_data = [
    # POM
    [1, "lib/view/pom/refil_user.dart", 52, "checkMesin(String scannedId)", "NfcCheckMesin", "idmesin", "String", "Check machine NFC"],
    [2, "lib/view/pom/refil_user.dart", 76, "updateStok(dynamic id, dynamic stok)", "savePomStock", "id, stock", "dynamic, dynamic", "Update POM stock"],
    [3, "lib/view/pom/refilpom.dart", 47, "updateStok(dynamic id, dynamic stok)", "savePomStock", "id, stock", "dynamic, dynamic", "Update POM stock"],
    
    # POM USER
    [4, "lib/view/pomuser/menu.dart", 78, "fetchPomLog()", "pomGetPomLog", "doctype", "String", "Fetch POM log"],
    [5, "lib/view/pomuser/menu.dart", 114, "fetchPomHistory()", "pomGetPomLog", "doctype", "String", "Fetch POM history"],
    [6, "lib/view/pomuser/menu.dart", 139, "fetchPomLogNew()", "pomGetPomLogNew", "doctype", "String", "Fetch POM log new"],
    [7, "lib/view/pomuser/menu.dart", 213, "savePomLogDone()", "pomSavePomLogDone", "id", "String?", "Save POM log done"],
    [8, "lib/view/pomuser/menu.dart", 281, "savePomLog(BuildContext context)", "pomSavePomLog", "id, qty, keterangan", "String, String, String", "Save POM log"],
    [9, "lib/view/pomuser/menu.dart", 503, "fetchGensetLog()", "pomGetPomLog", "doctype", "String", "Fetch genset log"],
    [10, "lib/view/pomuser/menu.dart", 535, "fetchGensetHistory()", "pomGetPomLog", "doctype", "String", "Fetch genset history"],
    [11, "lib/view/pomuser/menu.dart", 588, "startRequest()", "pomStartRequest", "pomId, qty", "String, String", "Start POM request"],
    [12, "lib/view/pomuser/menu.dart", 732, "saveGensetLog(BuildContext context)", "pomSavePomLog", "id, qty, keterangan", "String, String, String", "Save genset log"],
    
    # BS
    [13, "lib/view/bs/pilihjo.dart", 323, "loadJo()", "bsGetJO", "body", "Map<String, dynamic>", "Load JO"],
    [14, "lib/view/bs/hasilbs.dart", 80, "fetchBonSementaraData(String joList)", "bsGetAmount", "jo, jenis", "String, String", "Fetch BS amount"],
    [15, "lib/view/bs/pilihkapal.dart", 294, "loadVessel()", "bsGetVessel", "-", "-", "Load vessel"],
    [16, "lib/view/bs/pilihrealisasi.dart", 360, "loadRealisasi()", "bsGetReal", "jo, vessel", "String, String", "Load realizasi"],
    [17, "lib/view/bs/pilihto.dart", 459, "loadTo()", "bsGetTO", "jo, jenis, realisasi", "String, String, String", "Load TO"],
    [18, "lib/view/bs/prosespelindo.dart", 80, "fetchVendorOptions()", "bsGetDepoLuar", "-", "-", "Fetch vendor options"],
    [19, "lib/view/bs/prosespelindo.dart", 93, "fetchBonSementaraData()", "bsGetAmount", "-", "-", "Fetch bon sementara"],
    [20, "lib/view/bs/tipe-bs.dart", 39, "getUserStatus()", "bsCheckUser", "-", "-", "Get user status"],
    
    # AP
    [21, "lib/view/ap/location-ap.dart", 31, "getLocationData()", "apCheckUser", "-", "-", "Get location data"],
    [22, "lib/view/ap/pilihjo-ap.dart", 270, "loadJo()", "apGetJO", "body", "Map<String, dynamic>", "Load JO"],
    [23, "lib/view/ap/pilihdata-ap.dart", 606, "loadData()", "apGetData", "jo, jenis", "String, String", "Load data"],
    [24, "lib/view/ap/pilihvendor-ap.dart", 318, "loadVendor()", "apGetVendor", "jenis", "String", "Load vendor"],
    [25, "lib/view/ap/pilihto-ap.dart", 248, "loadTo()", "apGetData", "body", "Map<String, dynamic>", "Load TO"],
    [26, "lib/view/ap/pilihkapal-ap.dart", 241, "loadKapal()", "apGetData", "jo, jenis", "String, String", "Load kapal"],
    [27, "lib/view/ap/hasilap.dart", 217, "loadVendor()", "apGetVendor", "jenis, additionalBody", "String, Map", "Load vendor"],
    
    # ACCIDENT
    [28, "lib/view/accident_list.dart", 193, "fetchAccidents()", "taGetAccident", "accidentType", "String", "Fetch accidents"],
    [29, "lib/view/accident_list.dart", 306, "approve(...)", "taApprovalAccident", "id, action", "String, String", "Approve accident"],
    [30, "lib/view/accident_list.dart", 383, "delete(...)", "taDeleteAccident", "id", "String", "Delete accident"],
    [31, "lib/view/accident_list.dart", 483, "requestApproval(...)", "taRequestApprovalAccident", "id", "String", "Request approval"],
    [32, "lib/view/accident_container_list.dart", 112, "fetchAccidents()", "taGetAccident", "accidentType", "String", "Fetch container accidents"],
    [33, "lib/view/accident_container_list.dart", 222, "approve(...)", "taApprovalAccident", "id, action", "String, String", "Approve container accident"],
    [34, "lib/view/accident_container_list.dart", 332, "delete(...)", "taDeleteAccident", "id", "String", "Delete container accident"],
    [35, "lib/view/accident_container_list.dart", 376, "requestApproval(...)", "taRequestApprovalAccident", "id", "String", "Request container approval"],
    [36, "lib/view/accident_container.dart", 365, "_pickContainer()", "getContainer", "-", "-", "Pick container"],
    [37, "lib/view/accident_approval.dart", 88, "fetchAll()", "toGetAccident", "-", "-", "Fetch all accidents"],
    [38, "lib/view/accident_approval.dart", 114, "approve(Accident a)", "toApproveAccident", "id, approval", "String, String", "Approve accident"],
    [39, "lib/view/accident.dart", 412, "_pickDriver()", "getDriverName", "-", "-", "Pick driver"],
    [40, "lib/view/accident.dart", 503, "_pickTruck()", "getTruck", "-", "-", "Pick truck"],
    
    # SECURITY
    [41, "lib/view/security/form.dart", 64, "fetchDropdownData()", "getContainer", "-", "-", "Fetch dropdown data"],
    [42, "lib/view/security/form.dart", 295, "pickDriver(...)", "getDriverName", "-", "-", "Pick driver"],
    [43, "lib/view/security/form.dart", 387, "pickTruck(...)", "getTruck", "-", "-", "Pick truck"],
    [44, "lib/view/security/gatelog.dart", 237, "fetchGateLogs()", "sgGetGateLog", "-", "-", "Fetch gate logs"],
    [45, "lib/view/security/print.dart", 15, "fetchpdf(...)", "sgGetTO", "-", "-", "Fetch PDF"],
    [46, "lib/view/security/print.dart", 36, "printPdf(...)", "Dynamic", "id", "String", "Print PDF"],
    
    # VESSEL
    [47, "lib/view/vessel_schedule.dart", 36, "_fetchVesselNames()", "getVesselName", "-", "-", "Fetch vessel names"],
    [48, "lib/view/vessel_schedule.dart", 59, "(async)", "getPortOrigin", "vessel", "String", "Get port origin"],
    [49, "lib/view/vessel_schedule.dart", 64, "(async)", "getPortDestination", "vessel", "String", "Get port destination"],
    [50, "lib/view/vessel_schedule.dart", 118, "_fetchSchedule()", "getVesselSchedule", "port, eta", "String, String", "Fetch vessel schedule"],
    
    # SALES
    [51, "lib/view/sales_activity.dart", 69, "fetchSales()", "getSales", "-", "-", "Fetch sales"],
    [52, "lib/view/salescustomer.dart", 35, "customerKeychange(String id, String blacklist)", "saveCustomerBlacklist", "customerId, status", "String, String", "Save customer blacklist"],
    
    # PANTRY
    [53, "lib/view/pantry/pantry_orders.dart", 40, "_fetchBalance()", "pantryGetSaldo", "-", "-", "Fetch balance"],
    [54, "lib/view/pantry/pantry_orders.dart", 62, "_fetchHistory()", "pantryGetOrder", "-", "-", "Fetch history"],
    [55, "lib/view/pantry/pantry_cart.dart", 389, "checkout()", "pantryGetSaldo", "-", "-", "Checkout"],
    
    # PLUG
    [56, "lib/view/plug/menu.dart", 65, "fetchPlugLog()", "plugGetPlugLog", "doctype", "String", "Fetch plug log"],
    [57, "lib/view/plug/menu.dart", 98, "savePlugLog(...)", "plugSavePlugLog", "id, qty", "String, String", "Save plug log"],
    [58, "lib/view/plug/menu.dart", 158, "checkJo(String scanned)", "checkPlugJo", "docNo", "String", "Check JO"],
    
    # SIMULATE
    [59, "lib/view/simulate/pilihjo.dart", 367, "loadJo()", "bsSimulasiGetJO", "body", "Map<String, dynamic>", "Load simulated JO"],
    [60, "lib/view/simulate/pilihjo.dart", 439, "loadAmount()", "bsSimulasiGetAmount", "body", "Map<String, dynamic>", "Load simulated amount"],
    [61, "lib/view/simulate/simulatebs.dart", 53, "fetchLokasi()", "getLokasi", "statusAdaKantor", "String", "Fetch lokasi"],
    [62, "lib/view/simulate/simulatebs.dart", 69, "fetchVessel()", "getVessel", "-", "-", "Fetch vessel"],
    [63, "lib/view/simulate/prosespelindo.dart", 65, "fetchVendorOptions()", "bsGetDepoLuar", "-", "-", "Fetch vendor options"],
    [64, "lib/view/simulate/prosespelindo.dart", 103, "fetchBonSementaraData()", "bsSimulasiGetAmount", "-", "-", "Fetch bon sementara"],
    [65, "lib/view/simulate/simulatehasilbs.dart", 79, "_fetchBonSementara()", "bsSimulasiGetAmount", "-", "-", "Fetch bon sementara"],
    
    # DASHBOARD
    [66, "lib/view/dashboard_stripping_page.dart", 264, "updateStatus(...)", "dashboardStrippingUpdate", "id, status, containerNo", "String, String, String", "Update stripping status"],
    [67, "lib/view/dashboard_stuffing_page.dart", 277, "updateStatus(...)", "dashboardStuffingUpdate", "id, status, containerNo", "String, String, String", "Update stuffing status"],
    
    # DRIVER ORDER
    [68, "lib/view/driver_order_page.dart", 367, "(dynamic)", "Dynamic", "-", "-", "Dynamic driver order"],
    [69, "lib/view/driver_order_page.dart", 469, "saveJob(...)", "toSaveDriverJob", "-", "-", "Save driver job"],
    
    # OTHERS
    [70, "lib/view/widgets/card_layanan.dart", 1023, "checkNfc(String nfcId)", "checkNfc", "nfc", "String", "Check NFC"],
    [71, "lib/view/apiactionbutton.dart", 84, "fetchUsers()", "getSnituser", "-", "-", "Fetch users"],
    [72, "lib/view/usermap.dart", 49, "fetchUsers()", "getSnituser", "-", "-", "Fetch users"],
    
    # SERVICES
    [73, "lib/services/lifecycle_service.dart", 74, "_sendHeartbeat()", "saveUserHeartbeat", "lat, lng, accuracy", "String, String, double", "Send heartbeat"],
    [74, "lib/services/lifecycle_service.dart", 92, "_sendLifecyclePing(String status)", "saveUserHeartbeat", "lat, lng, accuracy", "String, String, double", "Send lifecycle ping"],
    [75, "lib/services/lifecycle_service.dart", 113, "_sendShutdownHeartbeat()", "killHeartbeat", "imei", "String", "Send shutdown heartbeat"],
]

for row_idx, row_data in enumerate(callapi_view_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

# ============== SHEET 3: Manual HTTP ==============
ws3 = wb.create_sheet("Manual_HTTP")

headers3 = ["No", "File", "Line", "Method", "Endpoint/Path", "Body Parameters", "Data Type"]
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

manual_http_data = [
    # CONTROLLER
    [1, "lib/controller/login_controller.dart", 109, "POST", "/api/token", "imei, version, timezone, latitude, longitude", "String, String, String, double?, double?"],
    [2, "lib/controller/login_controller.dart", 214, "POST", "apisnit", "modulCode: saveNfc, nfc", "String, String"],
    [3, "lib/controller/driver_manual_absen_controller.dart", 235, "POST", "/api/absensi", "type, imei, image, lat, lng, address", "String, String, String, double, double, String"],
    [4, "lib/controller/camera_controller.dart", 355, "POST", "/api/absensi", "Absensi data + image (base64)", "Various"],
    [5, "lib/controller/driver_absen_controller.dart", 355, "POST", "/api/absensi", "Absensi data + image (base64)", "Various"],
    [6, "lib/controller/manual_camera_controller.dart", 246, "POST", "/api/absensi", "Absensi data + image (base64)", "Various"],
    [7, "lib/controller/salesplan_controller.dart", 36, "POST", "apisnit", "Sales plan data", "Map<String, dynamic>"],
    [8, "lib/controller/salesplan_controller.dart", 190, "POST", "apisnit", "Sales plan data", "Map<String, dynamic>"],
    [9, "lib/controller/salesplan_controller.dart", 249, "POST", "apisnit", "Sales plan data", "Map<String, dynamic>"],
    [10, "lib/controller/salesplan_controller.dart", 382, "POST", "apisnit", "Sales plan data", "Map<String, dynamic>"],
    [11, "lib/controller/pantry_controller.dart", 94, "GET", "Local pantry server", "-", "-"],
    [12, "lib/controller/pantry_controller.dart", 144, "GET", "Local pantry server", "-", "-"],
    [13, "lib/controller/driver_map_controller.dart", 34, "GET", "Maps API", "-", "-"],
    [14, "lib/controller/driver_map_controller.dart", 71, "GET", "Maps API", "-", "-"],
    [15, "lib/controller/registration_controller.dart", 37, "POST", "Auth register", "Registration data", "Map<String, dynamic>"],
    [16, "lib/controller/registration_controller.dart", 38, "POST", "Auth register", "Registration data", "Map<String, dynamic>"],
    
    # VIEW
    [17, "lib/view/patroladmin.dart", 82, "GET", "Patrol admin API", "-", "-"],
    [18, "lib/view/patroladmin.dart", 95, "GET", "Patrol admin API", "-", "-"],
    [19, "lib/view/patroladmin.dart", 112, "POST", "Patrol admin API", "Patrol data", "Map<String, dynamic>"],
    [20, "lib/view/patroladmin.dart", 161, "DELETE", "Patrol admin API", "-", "-"],
    [21, "lib/view/patroladmin.dart", 189, "GET", "Patrol admin API", "-", "-"],
    [22, "lib/view/patroladmin.dart", 204, "POST", "Patrol admin API", "Patrol data", "Map<String, dynamic>"],
    [23, "lib/view/patroladmin.dart", 232, "DELETE", "Patrol admin API", "-", "-"],
    [24, "lib/view/accident.dart", 185, "GET", "Accident API", "-", "-"],
    [25, "lib/view/accident.dart", 343, "GET", "Accident API", "-", "-"],
    [26, "lib/view/accident_container.dart", 186, "GET", "Container API", "-", "-"],
    [27, "lib/view/accident_container.dart", 333, "GET", "Container API", "-", "-"],
    [28, "lib/view/driver_detail.dart", 124, "GET", "Driver API", "-", "-"],
    [29, "lib/view/driver_detail.dart", 171, "GET", "Driver API", "-", "-"],
    [30, "lib/view/address_route_page.dart", 29, "GET", "Address route", "-", "-"],
    [31, "lib/view/address_route_page.dart", 109, "GET", "Address route", "-", "-"],
    [32, "lib/view/sales_customer.dart", 46, "POST", "Customer API", "Customer data", "Map<String, dynamic>"],
    [33, "lib/view/sales_customer_detail.dart", 66, "POST", "Customer detail API", "Customer detail", "Map<String, dynamic>"],
    [34, "lib/view/salescustomer_detail.dart", 70, "POST", "Customer detail API", "Customer detail", "Map<String, dynamic>"],
    [35, "lib/view/inventorylist.dart", 54, "POST", "Inventory API", "Inventory data", "Map<String, dynamic>"],
    [36, "lib/view/trucklist.dart", 114, "POST", "Truck API", "Truck data", "Map<String, dynamic>"],
    [37, "lib/view/gensetlist.dart", 107, "POST", "Genset API", "Genset data", "Map<String, dynamic>"],
    [38, "lib/view/sasislist.dart", 107, "POST", "Sasis API", "Sasis data", "Map<String, dynamic>"],
    [39, "lib/view/bs/hasilbs.dart", 181, "POST", "BS amount API", "BS amount data", "Map<String, dynamic>"],
    [40, "lib/view/bs/hasilto.dart", 69, "POST", "BS TO API", "BS TO data", "Map<String, dynamic>"],
    [41, "lib/view/bs/hasilto.dart", 188, "POST", "BS TO API", "BS TO data", "Map<String, dynamic>"],
    [42, "lib/view/bs/pilihvendor.dart", 93, "POST", "Vendor API", "Vendor data", "Map<String, dynamic>"],
    [43, "lib/view/bs/pilihvendor.dart", 194, "POST", "Vendor API", "Vendor data", "Map<String, dynamic>"],
    [44, "lib/view/bs/pilihvendor.dart", 578, "POST", "Vendor API", "Vendor data", "Map<String, dynamic>"],
    [45, "lib/view/bs/hasilbs-coc.dart", 71, "POST", "BS COC API", "BS COC data", "Map<String, dynamic>"],
    [46, "lib/view/bs/hasilbs-coc.dart", 188, "POST", "BS COC API", "BS COC data", "Map<String, dynamic>"],
    [47, "lib/view/bs/pilih-coc.dart", 300, "POST", "BS COC API", "BS COC data", "Map<String, dynamic>"],
    [48, "lib/view/bs/pilih-coc-container.dart", 299, "POST", "COC container API", "COC container", "Map<String, dynamic>"],
    [49, "lib/view/bs/hasilrealisasi.dart", 327, "POST", "Realisasi API", "Realisasi data", "Map<String, dynamic>"],
    [50, "lib/view/ap/hasilap.dart", 564, "POST", "AP API", "AP data", "Map<String, dynamic>"],
    [51, "lib/view/pomuser/pommini.dart", 80, "POST", "POM API", "POM data", "Map<String, dynamic>"],
    [52, "lib/view/simulate/simulatehasilbs.dart", 116, "POST", "Simulate API", "Simulate data", "Map<String, dynamic>"],
    [53, "lib/view/security/form.dart", 267, "POST", "Form upload API", "Form upload", "Stream"],
    
    # MAIN
    [54, "lib/main.dart", 337, "POST", "Notifications API", "Notifications", "Map<String, dynamic>"],
    [55, "lib/main.dart", 700, "POST", "Update check API", "Update check", "Map<String, dynamic>"],
    
    # SERVICES
    [56, "lib/services/background_service.dart", 353, "POST", "Heartbeat API", "Heartbeat", "Map<String, dynamic>"],
]

for row_idx, row_data in enumerate(manual_http_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws3.cell(row=row_idx, column=col_idx, value=value)

# ============== SHEET 4: Summary ==============
ws4 = wb.create_sheet("Summary")

ws4.cell(row=1, column=1, value="API Documentation Summary")
ws4.cell(row=1, column=1).font = Font(bold=True, size=14)

ws4.cell(row=3, column=1, value="Category")
ws4.cell(row=3, column=2, value="Count")
ws4.cell(row=3, column=1).font = Font(bold=True)
ws4.cell(row=3, column=2).font = Font(bold=True)

summary_data = [
    ["callApiFetch (Controller)", 81],
    ["callApiFetch (View)", 75],
    ["callApiFetch (Services)", 3],
    ["Manual HTTP (Controller)", 16],
    ["Manual HTTP (View)", 37],
    ["Manual HTTP (Main)", 2],
    ["Manual HTTP (Services)", 1],
    ["", ""],
    ["TOTAL callApiFetch", 159],
    ["TOTAL Manual HTTP", 56],
    ["TOTAL API Connections", 215],
    ["", ""],
    ["Unique modulCode", 92],
]

for row_idx, row_data in enumerate(summary_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        ws4.cell(row=row_idx, column=col_idx, value=value)

# ============== Adjust Column Widths ==============
for ws in [ws1, ws2, ws3]:
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 40

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 15

# Save workbook
output_path = "/Users/user/.openclaw/workspace/TriDominic_API_Documentation.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")